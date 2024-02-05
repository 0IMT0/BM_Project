import pandas as pd
from openpyxl import load_workbook

def ve_to_offers(mel_df, offers_df):
    if 'mel_ve' not in offers_df.columns:
        merged_df = pd.merge(offers_df, mel_df[['bmu_id', 've']], left_on='id', right_on='bmu_id', how='left')
        merged_df.drop(columns=['bmu_id'], inplace=True)
        merged_df.rename(columns={'ve': 'mel_ve'}, inplace=True)
        return merged_df
    else:
        return offers_df

if __name__ == "__main__":
    excel_file_path = 'test.xlsx'

    try:
        mel_df = pd.read_excel(excel_file_path, sheet_name='mel')
        offers_df = pd.read_excel(excel_file_path, sheet_name='offers')
        updated_offers_df = ve_to_offers(mel_df, offers_df)

        # Load the workbook and find the position of the 'offers' sheet
        book = load_workbook(excel_file_path)
        sheet_order = book.sheetnames
        offers_position = sheet_order.index('offers') if 'offers' in sheet_order else len(sheet_order)

        # Remove the 'offers' sheet if it exists
        if 'offers' in book.sheetnames:
            del book['offers']

        # Convert the updated_offers_df DataFrame to an openpyxl worksheet
        from openpyxl.utils.dataframe import dataframe_to_rows
        ws = book.create_sheet('offers_temp')  # Temporarily create a new sheet

        for r_idx, row in enumerate(dataframe_to_rows(updated_offers_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Adjust the position of the new sheet to match the original 'offers' sheet
        sheet_names = book.sheetnames
        temp_index = sheet_names.index('offers_temp')
        book._sheets[temp_index], book._sheets[offers_position] = book._sheets[offers_position], book._sheets[temp_index]
        book._sheets[offers_position].title = 'offers'  # Rename the temp sheet to 'offers'

        # Save the workbook
        book.save(excel_file_path)
    except Exception as e:
        print(f"Error: {e}")
